"""
实现对excel文档的操作，包括:
1、加载已存在的xlsx文档并返回workbook对象
2、选择sheet页面：可通过加载激活sheet以及index,name选择sheet页面
"""
import os
import openpyxl


class OperateExcel():
    """
    具体实现xlsx文档的操作
    """
    def __init__(self, path, filename):
        self.path = path
        self.filename = filename
        self.wb = openpyxl.load_workbook(os.path.join(path, filename), data_only=True)
        self.sheet = None

    def get_sheet_by_index(self, index):
        """通过index索引选择sheet页"""
        self.sheet = self.wb.worksheets[index]

    def get_sheet_by_name(self, name):
        """通过页面名称选择sheet页"""
        self.sheet = self.wb[name]

    def get_active_sheet(self):
        """选择worksheet当前激活sheet页"""
        self.sheet = self.wb.active

    def enter_data_to_cell(self, row, column, data=''):
        """向cell(row,column)单元格写入数据"""
        self.sheet.cell(row=row, column=column).value = data

    def enter_data_to_cell_ex(self, row, column, data=''):
        """向指定单元格填入数据，并设置单元格格式"""
        pass

    def enter_datas_to_row(self, start, datas):
        """
        向行写入数据，指定开始坐标。
        :param start: 数据填充行的起始单元格坐标，可传入一个元组，(row,column)
        :param datas: 需要填充的数据，是一个数据列表
        :return:
        """
        for i, data in enumerate(datas):
            self.sheet.cell(row=start[0], column=start[1] + i).value = data

    def enter_datas_to_column(self, start, datas):
        """
        向列写入数据，指定开始坐标。
        :param start: 数据填充列的起始单元格坐标，可传入一个元组，(row,column)
        :param datas: 需要填充的数据，是一个数据列表
        :return:
        """
        for i, data in enumerate(datas):
            self.sheet.cell(row=start[0] + i, column=start[1]).value = data

    def get_cell_data(self, row, column):
        """
        获取指定单元格值
        :param row:
        :param column:
        :return: 返回指定单元格的值
        """
        return self.sheet.cell(row=row, column=column).value

    def get_row_datas(self, start, num):
        """
        获取从start开始，num个的数据，行
        :param start: 开始单元格坐标
        :param num: 获取数据数量
        :return: 返回数据列表
        """
        datas = []
        for i in range(0,num):
            datas.append(self.sheet.cell(row=start[0], column=start[1] + i).value)
        return datas

    def get_column_datas(self, start, num):
        """
        获取从start开始，num个的数据，列
        :param start: 开始单元格坐标
        :param num: 获取数据数量
        :return: 返回数据列表
        """
        datas = []
        for i in range(0, num):
            datas.append(self.sheet.cell(row=start[0] + i, column=start[1]).value)
        return datas

    def save(self):
        """保存文档(更新)"""
        self.wb.save(self.filename)

    def save_as(self, path, filename):
        """另存为文档"""
        self.wb.save(os.path.join(path, filename))

